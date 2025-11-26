# -*- coding: utf-8 -*-
import time
import sys
import os
import getpass
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

# Intentar importar dotenv (si está instalado)
try:
    from dotenv import load_dotenv
    DOTENV_DISPONIBLE = True
except ImportError:
    DOTENV_DISPONIBLE = False
    print("Advertencia: 'python-dotenv' no esta instalado. Se pediran credenciales manualmente.")
    print("Para instalarlo, usa: pip install python-dotenv")

# ============================================================================
# FUNCIONES AUXILIARES (Sin cambios)
# ============================================================================
def imprimir_error_seguro(mensaje, e):
    """Imprime un error sin fallar por caracteres especiales"""
    try:
        error_ascii = str(e).encode('ascii', 'ignore').decode('ascii')
        print(f"\n[ERROR] {mensaje}: {error_ascii}\n")
    except Exception:
        print(f"\n[ERROR] {mensaje}: Ocurrio un error no imprimible.")

def configurar_driver():
    """Configura el navegador Chrome con opciones optimizadas"""
    print("Configurando navegador Chrome...")
    chrome_options = Options()
    # chrome_options.add_argument('--headless')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument('--disable-blink-features=AutomationControlled')
    chrome_options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    chrome_options.add_experimental_option("prefs", {
        "profile.default_content_setting_values.notifications": 2
    })
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    print("Navegador configurado correctamente.\n")
    return driver

def scroll_pagina(driver, pausas=3):
    """Realiza scroll para cargar contenido dinámico"""
    altura_anterior = 0
    for i in range(pausas):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
        altura_nueva = driver.execute_script("return document.body.scrollHeight")
        if altura_nueva == altura_anterior: break
        altura_anterior = altura_nueva

def cerrar_modales(driver):
    """Intenta cerrar modales y pop-ups comunes"""
    selectores_cierre = [
        "//button[contains(text(), 'Entendido')]", "//button[contains(text(), 'Aceptar')]",
        "//button[contains(@class, 'close')]", "//button[contains(@aria-label, 'Close')]",
        "//button[contains(@aria-label, 'Cerrar')]", "button.ee-c-modal__close",
        "#onetrust-accept-btn-handler", "div#pop-alert button.close"
    ]
    for selector in selectores_cierre:
        try:
            if selector.startswith("//"): boton = driver.find_element(By.XPATH, selector)
            elif selector.startswith("#"): boton = driver.find_element(By.ID, selector.replace("#", ""))
            else: boton = driver.find_element(By.CSS_SELECTOR, selector)
            driver.execute_script("arguments[0].click();", boton)
            print(f"[OK] Modal cerrado: {selector[:30]}...")
            time.sleep(1)
        except: pass

# ============================================================================
# FUNCIONES DE LOGIN (Sin cambios en la lógica interna)
# ============================================================================
def login_elempleo(driver, email, password):
    """Login para Elempleo"""
    try:
        print("[...] Iniciando sesion en Elempleo...")
        driver.get("https://www.elempleo.com/co/iniciar-sesion")
        wait = WebDriverWait(driver, 15)
        email_field = wait.until(EC.presence_of_element_located((By.NAME, "email")))
        email_field.clear(); email_field.send_keys(email)
        password_field = driver.find_element(By.NAME, "password")
        password_field.clear(); password_field.send_keys(password)
        driver.find_element(By.XPATH, "//button[@type='submit']").click()
        time.sleep(5)
        if "iniciar-sesion" in driver.current_url:
            print("[ADVERTENCIA] Login en Elempleo no exitoso (revisa credenciales)")
            return False
        print("[OK] Login exitoso en Elempleo")
        return True
    except Exception as e:
        imprimir_error_seguro("Error en login de Elempleo", e)
        return False

def login_computrabajo(driver, email, password):
    """Login para Computrabajo"""
    try:
        print("[...] Iniciando sesion en Computrabajo...")
        driver.get("https://www.computrabajo.com.co/login")
        wait = WebDriverWait(driver, 15)
        email_field = wait.until(EC.presence_of_element_located((By.ID, "Email")))
        email_field.clear(); email_field.send_keys(email)
        password_field = driver.find_element(By.ID, "Password")
        password_field.clear(); password_field.send_keys(password)
        driver.find_element(By.ID, "btnIngresar").click()
        time.sleep(5)
        cerrar_modales(driver)
        print("[OK] Login completado en Computrabajo")
        return True
    except Exception as e:
        imprimir_error_seguro("Error en login de Computrabajo", e)
        return False

def login_magneto(driver, email, password):
    """Login para Magneto365"""
    try:
        print("[...] Iniciando sesion en Magneto365...")
        driver.get("https://www.magneto365.com/co/login")
        wait = WebDriverWait(driver, 15)
        email_field = wait.until(EC.presence_of_element_located((By.ID, "email")))
        email_field.clear(); email_field.send_keys(email)
        password_field = driver.find_element(By.ID, "password")
        password_field.clear(); password_field.send_keys(password)
        driver.find_element(By.XPATH, "//button[@type='submit']").click()
        time.sleep(5)
        cerrar_modales(driver)
        print("[OK] Login completado en Magneto365")
        return True
    except Exception as e:
        imprimir_error_seguro("Error en login de Magneto365", e)
        return False

# ============================================================================
# SCRAPERS PARA CADA PLATAFORMA (Lógica de scraping sin cambios)
# ============================================================================
# --- scraper_elempleo_robusto ---
def scraper_elempleo_robusto(puesto_buscado, email_usuario, clave_usuario, max_paginas=3):
    print("\n" + "="*60); print("INICIANDO SCRAPING EN ELEMPLEO.COM"); print("="*60)
    resultados = []; driver = None
    try:
        driver = configurar_driver()
        if not login_elempleo(driver, email_usuario, clave_usuario): print("[ADVERTENCIA] Continuando sin login...")
        termino_url = puesto_buscado.lower().replace(' ', '-')
        url_base = f"https://www.elempleo.com/co/ofertas-empleo/{termino_url}"
        for pagina in range(1, max_paginas + 1):
            print(f"\n[...] Procesando pagina {pagina}...")
            driver.get(url_base + (f"?page={pagina}" if pagina > 1 else ""))
            time.sleep(3); cerrar_modales(driver); scroll_pagina(driver, pausas=2)
            selectores = ["article.job-card-container", "div.job-item", "article[data-testid='job-card']"]
            ofertas = []; selector_usado = None
            for selector in selectores:
                try:
                    WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, selector)))
                    ofertas = driver.find_elements(By.CSS_SELECTOR, selector)
                    if ofertas: print(f"[OK] Ofertas encontradas con selector: {selector}"); selector_usado = selector; break
                except: continue
            if not ofertas: print(f"[ADVERTENCIA] No se encontraron ofertas en la pagina {pagina}"); break
            print(f"[...] Procesando {len(ofertas)} ofertas...")
            for idx, oferta in enumerate(ofertas, 1):
                try:
                    titulo_elem = oferta.find_element(By.CSS_SELECTOR, "a.job-title-link, a[data-testid='job-title']")
                    titulo = titulo_elem.text.strip(); enlace = titulo_elem.get_attribute('href')
                    try: empresa = oferta.find_element(By.CSS_SELECTOR, "span[data-e-company-name='true'], .company-name").text.strip()
                    except: empresa = "No especificada"
                    try: ubicacion = oferta.find_element(By.CSS_SELECTOR, "span[data-e-city-name='true'], .location").text.strip()
                    except: ubicacion = "No especificada"
                    try: fecha = oferta.find_element(By.CSS_SELECTOR, "span.job-date-info, .date").text.strip()
                    except: fecha = "No disponible"
                    try: salario = oferta.find_element(By.CSS_SELECTOR, "span.job-salary, .salary").text.strip()
                    except: salario = "No especificado"
                    print(f"  {idx}. {titulo[:50]}... | {empresa}")
                    resultados.append({'titulo': titulo, 'empresa': empresa, 'correo': "Ver en detalle", 'fecha': fecha, 'ciudad': ubicacion, 'salario': salario, 'tipo_contrato': "Ver detalle", 'enlace': enlace, 'fuente': 'ElEmpleo'})
                except Exception as e: print(f"  [ERROR] Error en oferta {idx}"); continue
            time.sleep(2)
    except Exception as e: imprimir_error_seguro("ERROR FATAL EN ELEMPLEO", e)
    finally:
        if driver: driver.quit()
    print(f"\n[OK] Total extraido de ElEmpleo: {len(resultados)} ofertas")
    return resultados

# --- scraper_computrabajo_robusto ---
def scraper_computrabajo_robusto(puesto_buscado, email_usuario, clave_usuario, max_paginas=3):
    print("\n" + "="*60); print("INICIANDO SCRAPING EN COMPUTRABAJO"); print("="*60)
    resultados = []; driver = None
    try:
        driver = configurar_driver()
        if not login_computrabajo(driver, email_usuario, clave_usuario): print("[ADVERTENCIA] Continuando sin login...")
        termino_url = puesto_buscado.lower().replace(' ', '-')
        url_base = f"https://www.computrabajo.com.co/trabajo-de-{termino_url}"
        for pagina in range(1, max_paginas + 1):
            url = f"{url_base}?p={pagina}" if pagina > 1 else url_base
            print(f"\n[...] Procesando pagina {pagina}...")
            driver.get(url)
            time.sleep(3); cerrar_modales(driver)
            selectores = ["article[class*='offer']", "div.bRS", "article.box_offer"]
            ofertas = []; selector_usado = None
            for selector in selectores:
                try:
                    WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, selector)))
                    ofertas = driver.find_elements(By.CSS_SELECTOR, selector)
                    if ofertas: print(f"[OK] Ofertas encontradas con selector: {selector}"); selector_usado = selector; break
                except: continue
            if not ofertas: print(f"[ADVERTENCIA] No se encontraron ofertas en la pagina {pagina}"); break
            print(f"[...] Procesando {len(ofertas)} ofertas...")
            for idx, oferta in enumerate(ofertas, 1):
                try:
                    titulo_elem = oferta.find_element(By.CSS_SELECTOR, "a.js-o-link, h2 a")
                    titulo = titulo_elem.text.strip(); enlace = titulo_elem.get_attribute('href')
                    try: empresa = oferta.find_element(By.CSS_SELECTOR, "p a.it-blank, .company").text.strip()
                    except: empresa = "No especificada"
                    try: ubicacion = oferta.find_element(By.CSS_SELECTOR, "span.list-char-item, .location").text.strip()
                    except: ubicacion = "No especificada"
                    try: fecha_elem = oferta.find_element(By.CSS_SELECTOR, "p.fs16.fc_aux, .date"); fecha = fecha_elem.text.strip().split('\n')[-1]
                    except: fecha = "No disponible"
                    print(f"  {idx}. {titulo[:50]}... | {empresa}")
                    resultados.append({'titulo': titulo, 'empresa': empresa, 'correo': "Ver en detalle", 'fecha': fecha, 'ciudad': ubicacion, 'salario': "Ver detalle", 'tipo_contrato': "Ver detalle", 'enlace': enlace, 'fuente': 'Computrabajo'})
                except Exception as e: print(f"  [ERROR] Error en oferta {idx}"); continue
            time.sleep(2)
    except Exception as e: imprimir_error_seguro("ERROR FATAL EN COMPUTRABAJO", e)
    finally:
        if driver: driver.quit()
    print(f"\n[OK] Total extraido de Computrabajo: {len(resultados)} ofertas")
    return resultados

# --- scraper_magneto_robusto ---
def scraper_magneto_robusto(puesto_buscado, email_usuario, clave_usuario, max_paginas=3):
    print("\n" + "="*60); print("INICIANDO SCRAPING EN MAGNETO365"); print("="*60)
    resultados = []; driver = None
    try:
        driver = configurar_driver()
        if not login_magneto(driver, email_usuario, clave_usuario): print("[ADVERTENCIA] Continuando sin login...")
        termino_url = puesto_buscado.lower().replace(' ', '+')
        url_base = f"https://www.magneto365.com/co/empleos?q={termino_url}"
        print(f"\n[...] Accediendo a busqueda..."); driver.get(url_base); time.sleep(5); cerrar_modales(driver)
        for intento in range(max_paginas):
            selectores = ["div[class*='card-job']", "div.job-card", "article.job-item"]
            ofertas = []; selector_usado = None
            for selector in selectores:
                try:
                    WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, selector)))
                    ofertas = driver.find_elements(By.CSS_SELECTOR, selector)
                    if ofertas: selector_usado = selector; break
                except: continue
            if not ofertas: print(f"[ADVERTENCIA] No se encontraron mas ofertas (intento {intento + 1})"); break
            print(f"\n[...] Procesando {len(ofertas)} ofertas...")
            ofertas_nuevas = 0
            for idx, oferta in enumerate(ofertas, 1):
                try:
                    titulo_elem = oferta.find_element(By.CSS_SELECTOR, "h2.card-job-title, h3 a")
                    titulo = titulo_elem.text.strip()
                    enlace_elem = titulo_elem.find_element(By.TAG_NAME, "a") if titulo_elem.tag_name != "a" else titulo_elem
                    enlace = enlace_elem.get_attribute('href')
                    if any(r['enlace'] == enlace for r in resultados): continue
                    try: empresa = oferta.find_element(By.CSS_SELECTOR, "p.card-job-company, .company").text.strip()
                    except: empresa = "No especificada"
                    try: ubicacion = oferta.find_element(By.CSS_SELECTOR, "p.card-job-location, .location").text.strip()
                    except: ubicacion = "No especificada"
                    try: fecha = oferta.find_element(By.CSS_SELECTOR, "p.card-job-date, .date").text.strip()
                    except: fecha = "No disponible"
                    print(f"  {len(resultados) + 1}. {titulo[:50]}... | {empresa}")
                    ofertas_nuevas += 1
                    resultados.append({'titulo': titulo, 'empresa': empresa, 'correo': "Ver en detalle", 'fecha': fecha, 'ciudad': ubicacion, 'salario': "Ver detalle", 'tipo_contrato': "Ver detalle", 'enlace': enlace, 'fuente': 'Magneto365'})
                except Exception as e: continue
            if ofertas_nuevas == 0: print("[ADVERTENCIA] No se encontraron nuevas ofertas"); break
            if intento < max_paginas - 1: print("[...] Cargando mas ofertas..."); scroll_pagina(driver, pausas=4); time.sleep(3)
    except Exception as e: imprimir_error_seguro("ERROR FATAL EN MAGNETO", e)
    finally:
        if driver: driver.quit()
    print(f"\n[OK] Total extraido de Magneto: {len(resultados)} ofertas")
    return resultados

# ============================================================================
# GUARDAR EN EXCEL (Sin cambios)
# ============================================================================
def guardar_excel_mejorado(ofertas, puesto_buscado):
    if not ofertas: print("\n[ADVERTENCIA] No se encontraron ofertas para guardar."); return
    wb = Workbook(); ws = wb.active; ws.title = "Ofertas Laborales"
    encabezados = ["#", "Titulo", "Empresa", "Correo/Contacto", "Fecha", "Ciudad", "Salario", "Tipo Contrato", "Enlace", "Fuente"]
    ws.append(encabezados)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]: cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal="center", vertical="center")
    for idx, oferta in enumerate(ofertas, 1):
        ws.append([idx, oferta['titulo'], oferta['empresa'], oferta['correo'], oferta['fecha'], oferta['ciudad'], oferta.get('salario', 'No especificado'), oferta['tipo_contrato'], oferta['enlace'], oferta['fuente']])
    dims = {'A': 5, 'B': 50, 'C': 30, 'D': 30, 'E': 20, 'F': 20, 'G': 20, 'H': 20, 'I': 60, 'J': 15}
    for col, width in dims.items(): ws.column_dimensions[col].width = width
    nombre_puesto_limpio = "".join(c for c in puesto_buscado if c.isalnum() or c in (' ', '-')).strip()
    nombre_archivo = f"Ofertas_{nombre_puesto_limpio}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    try:
        wb.save(nombre_archivo)
        print("\n" + "="*60); print(f"[OK] EXITO! {len(ofertas)} ofertas guardadas en:"); print(f"  Archivo: {nombre_archivo}"); print("="*60)
    except PermissionError:
        print("\n" + "!"*60); print(f"[ERROR] No se pudo guardar '{nombre_archivo}'"); print("  El archivo esta abierto. Cierrelo e intente nuevamente."); print("!"*60)
    except Exception as e: imprimir_error_seguro("ERROR AL GUARDAR EXCEL", e)

# ============================================================================
# FUNCIÓN PRINCIPAL (MODIFICADA PARA LEER CLAVES ESPECÍFICAS)
# ============================================================================
def main():
    print("\n" + "="*60); print("BUSCADOR AUTOMATICO DE OFERTAS LABORALES"); print("Version - Claves Separadas"); print("="*60)
    puesto_buscado = "Ingeniero Mecanico"
    
    # --- SISTEMA DE CREDENCIALES MEJORADO ---
    email_usuario = None
    claves = {'elempleo': None, 'computrabajo': None, 'magneto': None} # Diccionario para claves
    
    # OPCIÓN 1: Archivo .env
    if DOTENV_DISPONIBLE:
        dotenv_path = os.path.join(os.path.dirname(__file__), '.env')
        if os.path.exists(dotenv_path):
            load_dotenv(dotenv_path=dotenv_path)
            email_usuario = os.getenv('EMAIL_EMPLEO')
            # --- CAMBIO CLAVE: Leer claves específicas ---
            claves['elempleo'] = os.getenv('PASSWORD_ELEMpleo')
            claves['computrabajo'] = os.getenv('PASSWORD_COMPUTRABAJO')
            claves['magneto'] = os.getenv('PASSWORD_MAGNETO')
            
            # Verificar si se cargaron todas
            if email_usuario and all(claves.values()):
                print("\n[OK] Credenciales cargadas desde archivo .env")
            else: # Si falta alguna clave en .env, anular todo para pedir manual
                print("[INFO] Faltan claves especificas en .env. Se pediran manualmente.")
                email_usuario = None # Forzar pedir manual si falta algo
                claves = {'elempleo': None, 'computrabajo': None, 'magneto': None}
        else:
             print("\n[INFO] Archivo .env no encontrado.")

    # OPCIÓN 2: Variables de sistema (menos probable que estén las 3 claves)
    if not email_usuario or not all(claves.values()):
        email_usuario = os.getenv('EMAIL_EMPLEO')
        claves['elempleo'] = os.getenv('PASSWORD_ELEMpleo')
        claves['computrabajo'] = os.getenv('PASSWORD_COMPUTRABAJO')
        claves['magneto'] = os.getenv('PASSWORD_MAGNETO')
        
        if email_usuario and all(claves.values()):
            print("\n[OK] Credenciales cargadas desde variables de entorno del sistema")
        else: # Si falta algo, anular para pedir manual
             email_usuario = None
             claves = {'elempleo': None, 'computrabajo': None, 'magneto': None}

    # OPCIÓN 3: Pedir manualmente
    if not email_usuario or not all(claves.values()):
        print("\n[ADVERTENCIA] No se encontraron todas las credenciales configuradas.")
        print("Se pediran manualmente (recomendado usar archivo .env).")
        print("Asegurate de crear un archivo .env con:")
        print("   EMAIL_EMPLEO=tu_email@ejemplo.com")
        print("   PASSWORD_ELEMpleo=tu_clave_elempleo")
        print("   PASSWORD_COMPUTRABAJO=tu_clave_computrabajo")
        print("   PASSWORD_MAGNETO=tu_clave_magneto")
        print("-" * 60)
        
        email_usuario = input("\nIngresa tu email (usado en las 3 plataformas): ")
        if email_usuario:
             # Pedir cada clave de forma segura
             claves['elempleo'] = getpass.getpass("Ingresa tu contrasena para Elempleo (no se vera): ")
             claves['computrabajo'] = getpass.getpass("Ingresa tu contrasena para Computrabajo (no se vera): ")
             claves['magneto'] = getpass.getpass("Ingresa tu contrasena para Magneto365 (no se vera): ")
        else: # Si no ingresa email, no tiene sentido pedir claves
            claves = {'elempleo': None, 'computrabajo': None, 'magneto': None}

    # Verificar final
    if not email_usuario or not all(claves.values()):
        print("\n[ERROR] Error: Se requieren el email y las 3 contrasenas para continuar.")
        print("Presione ENTER para salir...")
        sys.stdin.read(1)
        return
    
    print(f"\nUsuario: {email_usuario}")
    print(f"Buscando: '{puesto_buscado}'")
    print("Este proceso puede tardar varios minutos...\n")
    
    todas_ofertas = []
    
    # --- Ejecutar scrapers pasando la clave CORRECTA a cada uno ---
    todas_ofertas.extend(scraper_elempleo_robusto(puesto_buscado, email_usuario, claves['elempleo'], max_paginas=2))
    todas_ofertas.extend(scraper_computrabajo_robusto(puesto_buscado, email_usuario, claves['computrabajo'], max_paginas=2))
    todas_ofertas.extend(scraper_magneto_robusto(puesto_buscado, email_usuario, claves['magneto'], max_paginas=2))
    
    # Guardar resultados
    if todas_ofertas:
        guardar_excel_mejorado(todas_ofertas, puesto_buscado)
    else:
        print("\n[ADVERTENCIA] No se encontraron ofertas en ninguna plataforma.")
    
    print("\n[OK] Proceso completado.")
    print("Presione ENTER para salir...")
    sys.stdin.read(1)

if __name__ == "__main__":
    main()