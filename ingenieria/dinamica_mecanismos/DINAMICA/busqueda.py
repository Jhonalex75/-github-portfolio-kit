# -*- coding: utf-8 -*-
import time
import re
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.keys import Keys

# ============================================================================
# FUNCION UTILIDAD PARA ERRORES
# ============================================================================


def imprimir_error_seguro(mensaje, e):
    """ Imprime un error sin fallar por caracteres especiales """
    try:
        error_ascii = str(e).encode('ascii', 'ignore').decode('ascii')
        print(f"\n{mensaje}: {error_ascii}\n")
    except Exception:
        print(f"\n{mensaje}: Ocurrio un error no imprimible.")

# ============================================================================
# CONFIGURACION Y FUNCIONES AUXILIARES
# ============================================================================

def configurar_driver():
    """Configura el navegador Chrome con opciones optimizadas"""
    print("Configurando navegador Chrome...")
    chrome_options = Options()
    # Descomenta la siguiente linea para que NO se abra la ventana de Chrome
    # chrome_options.add_argument('--headless') 
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument('--disable-blink-features=AutomationControlled')
    chrome_options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36')
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    print("Navegador configurado correctamente.\n")
    return driver

def scroll_pagina(driver, pausas=3):
    """Realiza scroll para cargar contenido dinamico"""
    for _ in range(pausas):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)

# ============================================================================
# SCRAPER PARA ELEMPLEO.COM
# ============================================================================
def scraper_elempleo_robusto(puesto_buscado, max_paginas=3):
    print("\n" + "="*60)
    print("INICIANDO SCRAPING EN ELEMPLEO.COM (Version Robusta)")
    print("="*60)
    
    resultados = []
    driver = None 
    
    try:
        driver = configurar_driver()
        termino_url = puesto_buscado.lower().replace(' ', '-')
        url_base = f"https://www.elempleo.com/co/ofertas-empleo/{termino_url}"
        
        for pagina in range(1, max_paginas + 1):
            url = f"{url_base}?page={pagina}" if pagina > 1 else url_base
            print(f"\nProcesando pagina {pagina}...")
            print(f"URL: {url}")
            driver.get(url)
            
            # --- ¡¡CORRECCION CLAVE!! ---
            # Esperamos 3 segundos a que aparezcan TODOS los pop-ups
            print("Esperando pop-ups y banners de cookies...")
            time.sleep(3) 
            
            # 1. Clic en "Entendido" del banner de Cookies (EL PROBLEMA REAL)
            try:
                # Usamos XPath para buscar un boton que contenga el texto "Entendido"
                cookie_button = driver.find_element(By.XPATH, "//button[contains(text(), 'Entendido')]")
                cookie_button.click()
                print("Banner de cookies 'Entendido' cerrado.")
                time.sleep(1)
            except Exception:
                print("No se encontro banner de cookies.")
                
            # 2. Clic en "X" del modal "Destaca tu HV"
            try:
                modal_close_button = driver.find_element(By.CSS_SELECTOR, "button.ee-c-modal__close")
                if modal_close_button:
                    modal_close_button.click()
                    print("Cerrando modal de 'Destaca tu HV'.")
                    time.sleep(1)
            except Exception:
                print("No se encontro modal de pop-up 'Destaca HV'.")
            
            scroll_pagina(driver, pausas=2)
            SELECTOR_CONTENEDOR_OFERTA = "article.job-card-container" 
            
            try:
                WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, SELECTOR_CONTENEDOR_OFERTA))
                )
                print("Ofertas cargadas correctamente!")
            except TimeoutException:
                print(f"TIMEOUT: No se encontraron ofertas en la pagina {pagina} (Selector: {SELECTOR_CONTENEDOR_OFERTA})")
                break
            
            ofertas = driver.find_elements(By.CSS_SELECTOR, SELECTOR_CONTENEDOR_OFERTA)
            print(f"Se encontraron {len(ofertas)} ofertas en la pagina {pagina}")
            
            for idx, oferta in enumerate(ofertas, 1):
                try:
                    titulo_elem = oferta.find_element(By.CSS_SELECTOR, "a.job-title-link")
                    titulo = titulo_elem.text.strip()
                    enlace = titulo_elem.get_attribute('href')
                    empresa = oferta.find_element(By.CSS_SELECTOR, "span[data-e-company-name='true']").text.strip()
                    ubicacion = oferta.find_element(By.CSS_SELECTOR, "span[data-e-city-name='true']").text.strip()
                    fecha = oferta.find_element(By.CSS_SELECTOR, "span.job-date-info").text.strip()
                    
                    try:
                        salario = oferta.find_element(By.CSS_SELECTOR, "span.job-salary").text.strip()
                    except:
                        salario = "No especificado"
                    
                    print(f"  + {idx}. {titulo[:50]}... | {empresa}")
                    
                    resultados.append({
                        'titulo': titulo, 'empresa': empresa, 'correo': "Ver en detalle",
                        'fecha': fecha, 'ciudad': ubicacion, 'salario': salario,
                        'tipo_contrato': "Ver detalle", 'enlace': enlace, 'fuente': 'ElEmpleo'
                    })
                except Exception as e:
                    imprimir_error_seguro(f"  - Oferta {idx}: Error al leer datos", e)
                    continue
            time.sleep(2)
    except Exception as e:
        imprimir_error_seguro("ERROR FATAL EN ELEMPLEO", e)
    finally:
        if driver:
            driver.quit()
    print(f"\nTotal extraido de ElEmpleo: {len(resultados)} ofertas")
    return resultados

# ============================================================================
# SCRAPER PARA MAGNETO365
# ============================================================================
def scraper_magneto_robusto(puesto_buscado, max_paginas=3):
    print("\n" + "="*60)
    print("INICIANDO SCRAPING EN MAGNETO365 (Version Robusta)")
    print("="*60)
    
    resultados = []
    driver = None
    
    try:
        driver = configurar_driver()
        termino_url = puesto_buscado.lower().replace(' ', '+')
        url_base = f"https://www.magneto365.com/co/empleos?q={termino_url}"
        
        for pagina in range(1, max_paginas + 1):
            if pagina > 1:
                print("Magneto usa scroll infinito, haciendo mas scroll...")
                scroll_pagina(driver, pausas=4)
            else:
                print(f"\nProcesando pagina 1 (y siguientes por scroll)...")
                print(f"URL: {url_base}")
                driver.get(url_base)
                print("Esperando pop-ups y banners de cookies...")
                time.sleep(5)
            
            # --- MANEJO DE COOKIES EN MAGNETO ---
            if pagina == 1: # Solo hacerlo la primera vez
                try:
                    cookie_button = driver.find_element(By.XPATH, "//button[contains(text(), 'Aceptar')]")
                    cookie_button.click()
                    print("Banner de cookies 'Aceptar' cerrado.")
                    time.sleep(1)
                except Exception:
                    print("No se encontro banner de cookies en Magneto.")

            SELECTOR_CONTENEDOR_OFERTA = "div[class*='card-job']" 
            try:
                WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, SELECTOR_CONTENEDOR_OFERTA))
                )
                print("Ofertas cargadas correctamente!")
            except TimeoutException:
                print(f"TIMEOUT: No se encontraron ofertas (Selector: {SELECTOR_CONTENEDOR_OFERTA})")
                break
            
            ofertas = driver.find_elements(By.CSS_SELECTOR, SELECTOR_CONTENEDOR_OFERTA)
            print(f"Se encontraron {len(ofertas)} ofertas cargadas hasta ahora.")
            
            nuevas_ofertas_encontradas = 0
            for idx, oferta in enumerate(ofertas, 1):
                try:
                    titulo_elem = oferta.find_element(By.CSS_SELECTOR, "h2.card-job-title")
                    titulo = titulo_elem.text.strip()
                    enlace = titulo_elem.find_element(By.TAG_NAME, "a").get_attribute('href')
                    
                    if any(r['enlace'] == enlace for r in resultados):
                        continue

                    empresa = oferta.find_element(By.CSS_SELECTOR, "p.card-job-company").text.strip()
                    ubicacion = oferta.find_element(By.CSS_SELECTOR, "p.card-job-location").text.strip()
                    fecha = oferta.find_element(By.CSS_SELECTOR, "p.card-job-date").text.strip()

                    print(f"  + {idx}. {titulo[:50]}... | {empresa}")
                    nuevas_ofertas_encontradas += 1
                    
                    resultados.append({
                        'titulo': titulo, 'empresa': empresa, 'correo': "Ver en detalle",
                        'fecha': fecha, 'ciudad': ubicacion, 'salario': "Ver detalle",
                        'tipo_contrato': "Ver detalle", 'enlace': enlace, 'fuente': 'Magneto365'
                    })
                except Exception as e:
                    imprimir_error_seguro(f"  - Oferta {idx}: Error al leer datos", e)
                    continue
            
            if nuevas_ofertas_encontradas == 0 and pagina > 1:
                print("No se cargaron mas ofertas con el scroll. Terminando Magneto.")
                break
            if pagina == max_paginas:
                break
    except Exception as e:
        imprimir_error_seguro("ERROR FATAL EN MAGNETO", e)
    finally:
        if driver:
            driver.quit()
    print(f"\nTotal extraido de Magneto: {len(resultados)} ofertas")
    return resultados

# ============================================================================
# SCRAPER PARA COMPUTRABAJO
# ============================================================================
def scraper_computrabajo_robusto(puesto_buscado, max_paginas=3):
    print("\n" + "="*60)
    print("INICIANDO SCRAPING EN COMPUTRABAJO (Version Robusta)")
    print("="*60)
    
    resultados = []
    driver = None
    
    try:
        driver = configurar_driver()
        termino_url = puesto_buscado.lower().replace(' ', '-')
        url_base = f"https://www.computrabajo.com.co/trabajo-de-{termino_url}"
        
        for pagina in range(1, max_paginas + 1):
            url = f"{url_base}?p={pagina}" if pagina > 1 else url_base
            print(f"\nProcesando pagina {pagina}...")
            print(f"URL: {url}")
            driver.get(url)
            
            print("Esperando pop-ups y banners de cookies...")
            time.sleep(2)
            
            # 1. MANEJO DE COOKIES EN COMPUTRABAJO
            try:
                # El boton tiene id 'onetrust-accept-btn-handler'
                cookie_button = driver.find_element(By.ID, "onetrust-accept-btn-handler")
                cookie_button.click()
                print("Banner de cookies 'Aceptar' cerrado.")
                time.sleep(1)
            except Exception:
                print("No se encontro banner de cookies en Computrabajo.")

            # 2. MANEJO DE POP-UP DE ALERTA
            try:
                close_button = driver.find_element(By.CSS_SELECTOR, "div#pop-alert button.close")
                if close_button:
                    close_button.click()
                    print("Cerrando pop-up de 'Alerta de Empleo'.")
                    time.sleep(1)
            except Exception:
                print("No se encontro pop-up de alerta.")
                
            SELECTOR_CONTENEDOR_OFERTA = "article[class*='offer']"
            
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, SELECTOR_CONTENEDOR_OFERTA))
                )
                print("Ofertas cargadas correctamente!")
            except TimeoutException:
                print(f"TIMEOUT: No se encontraron ofertas en la pagina {pagina} (Selector: {SELECTOR_CONTENEDOR_OFERTA})")
                break
                
            ofertas = driver.find_elements(By.CSS_SELECTOR, SELECTOR_CONTENEDOR_OFERTA)
            print(f"Se encontraron {len(ofertas)} ofertas en la pagina {pagina}")
            
            for idx, oferta in enumerate(ofertas, 1):
                try:
                    titulo_elem = oferta.find_element(By.CSS_SELECTOR, "a.js-o-link")
                    titulo = titulo_elem.text.strip()
                    enlace = titulo_elem.get_attribute('href')
                    empresa = oferta.find_element(By.CSS_SELECTOR, "p a.it-blank").text.strip()
                    ubicacion = oferta.find_element(By.CSS_SELECTOR, "span.list-char-item").text.strip()
                    
                    try:
                        fecha = oferta.find_element(By.CSS_SELECTOR, "p.fs16.fc_aux").text.strip().split('\n')[-1]
                    except:
                        fecha = "No disponible"
                    
                    print(f"  + {idx}. {titulo[:50]}... | {empresa}")
                    
                    resultados.append({
                        'titulo': titulo, 'empresa': empresa, 'correo': "Ver en detalle",
                        'fecha': fecha, 'ciudad': ubicacion, 'salario': "Ver detalle",
                        'tipo_contrato': "Ver detalle", 'enlace': enlace, 'fuente': 'Computrabajo'
                    })
                except Exception as e:
                    imprimir_error_seguro(f"  - Oferta {idx}: Error al leer datos", e)
                    continue
    except Exception as e:
        imprimir_error_seguro("ERROR FATAL EN COMPUTRABAJO", e)
    finally:
        if driver:
            driver.quit()
    print(f"\nTotal extraido de Computrabajo: {len(resultados)} ofertas")
    return resultados


# ============================================================================
# FUNCION PARA GUARDAR EN EXCEL
# ============================================================================
def guardar_excel_mejorado(ofertas, puesto_buscado):
    if not ofertas:
        print("\nNo se encontraron ofertas para guardar.")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Ofertas Laborales"
    
    encabezados = ["#", "Titulo", "Empresa", "Correo/Contacto", "Fecha", "Ciudad", "Salario", "Tipo Contrato", "Enlace", "Fuente"]
    ws.append(encabezados)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for idx, oferta in enumerate(ofertas, 1):
        ws.append([
            idx, oferta['titulo'], oferta['empresa'], oferta['correo'],
            oferta['fecha'], oferta['ciudad'], oferta.get('salario', 'No especificado'),
            oferta['tipo_contrato'], oferta['enlace'], oferta['fuente']
        ])
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 60
    ws.column_dimensions['J'].width = 15
    
    nombre_puesto_limpio = "".join(c for c in puesto_buscado if c.isalnum() or c in (' ', '-')).strip()
    nombre_archivo = f"Ofertas_{nombre_puesto_limpio}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    try:
        wb.save(nombre_archivo)
        print("\n" + "="*60)
        print(f"EXITO! {len(ofertas)} ofertas guardadas en:")
        print(f"Archivo: {nombre_archivo}")
        print("="*60)
    except PermissionError:
        print("\n" + "!"*60)
        print(f"ERROR: No se pudo guardar el archivo '{nombre_archivo}'.")
        print("El archivo Excel esta abierto. Por favor, cierrelo y vuelva a intentarlo.")
        print("!"*60)
    except Exception as e:
        imprimir_error_seguro("ERROR AL GUARDAR EXCEL", e)


# ============================================================================
# FUNCION PRINCIPAL
# ============================================================================
def main():
    print("\n" + "="*60)
    print("BUSCADOR AUTOMATICO DE OFERTAS LABORALES (V. Final)")
    print("="*60)
    
    # --- Busqueda fijada para eliminar error de input() ---
    puesto_buscado = "Ingeniero Mecanico"
    
    print(f"\nBuscando ofertas para: '{puesto_buscado}'")
    print("Este proceso puede tardar varios minutos...\n")
    
    todas_ofertas = []
    
    # --- Ejecutar scrapers ---
    todas_ofertas.extend(scraper_elempleo_robusto(puesto_buscado, max_paginas=2))
    todas_ofertas.extend(scraper_computrabajo_robusto(puesto_buscado, max_paginas=2))
    todas_ofertas.extend(scraper_magneto_robusto(puesto_buscado, max_paginas=2))
    
    # --- Guardar resultados ---
    if todas_ofertas:
        guardar_excel_mejorado(todas_ofertas, puesto_buscado)
    else:
        print("\nNo se encontraron ofertas en ninguna plataforma.")
    
    print("\nProceso completado.")
    print("Presiona ENTER para salir...")
    sys.stdin.read(1) # Pausa segura que no usa input()

if __name__ == "__main__":
    main()